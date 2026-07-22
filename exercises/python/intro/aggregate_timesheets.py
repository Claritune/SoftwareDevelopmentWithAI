#!/usr/bin/env python3
"""Aggregate weekly team timesheets into a combined monthly report."""

from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

EXPECTED_COLUMNS = [
    "Employee Name",
    "Role",
    "Project",
    "Mon",
    "Tue",
    "Wed",
    "Thu",
    "Fri",
    "Hour Type",
]

REPORT_CATEGORIES = ["Development", "QA Testing", "Leave", "Training", "Meetings"]

DEFAULT_CATEGORY_MAPPING = {
    "Development": "Development",
    "QA Testing": "QA Testing",
    "Vacation": "Leave",
    "Sick Day": "Leave",
    "Course": "Training",
    "Meetings": "Meetings",
}

NO_PROJECT_LABEL = "(No Project)"
OUTPUT_FILENAME = "Monthly_Report.xlsx"
EXCLUDED_PATTERNS = ("Monthly_Report", "REFERENCE", "Template")


class TimesheetError(Exception):
    """Raised when input validation or processing fails."""


def load_category_mapping(template_path: Path | None) -> dict[str, str]:
    """Load category mapping from template, or use defaults."""
    if template_path is None or not template_path.exists():
        return DEFAULT_CATEGORY_MAPPING.copy()

    wb = openpyxl.load_workbook(template_path, data_only=True)
    if "Category Mapping" not in wb.sheetnames:
        return DEFAULT_CATEGORY_MAPPING.copy()

    ws = wb["Category Mapping"]
    mapping: dict[str, str] = {}
    for row in range(2, ws.max_row + 1):
        source = ws.cell(row=row, column=1).value
        target = ws.cell(row=row, column=2).value
        if source and target:
            mapping[str(source)] = str(target)

    return mapping or DEFAULT_CATEGORY_MAPPING.copy()


def is_timesheet_file(path: Path) -> bool:
    """Return True if the file looks like a team timesheet input."""
    name = path.name
    if any(pattern in name for pattern in EXCLUDED_PATTERNS):
        return False
    if name == OUTPUT_FILENAME:
        return False
    return path.suffix.lower() == ".xlsx"


def validate_sheet_headers(ws) -> None:
    """Ensure the sheet has the expected column layout."""
    headers = [ws.cell(row=1, column=col).value for col in range(1, len(EXPECTED_COLUMNS) + 1)]
    if headers != EXPECTED_COLUMNS:
        raise TimesheetError(
            f"Sheet '{ws.title}' has unexpected columns.\n"
            f"  Expected: {EXPECTED_COLUMNS}\n"
            f"  Found:    {headers}"
        )


def row_total_hours(ws, row: int) -> float:
    """Sum Mon–Fri hours for a row, treating blanks as zero."""
    total = 0.0
    for col in range(4, 9):
        value = ws.cell(row=row, column=col).value
        if value is None:
            continue
        try:
            total += float(value)
        except (TypeError, ValueError) as exc:
            raise TimesheetError(
                f"Invalid hour value in sheet '{ws.title}', row {row}: {value!r}"
            ) from exc
    return total


def read_timesheet_file(path: Path, category_mapping: dict[str, str]) -> tuple[dict, dict, int]:
    """
    Read one team workbook and return project/role aggregations plus row count.

    Returns:
        (by_project, by_role, rows_processed)
        Each aggregation dict maps key -> category -> hours.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    by_project: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    by_role: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    rows_processed = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        validate_sheet_headers(ws)

        for row in range(2, ws.max_row + 1):
            employee = ws.cell(row=row, column=1).value
            if not employee:
                continue

            role = ws.cell(row=row, column=2).value
            project = ws.cell(row=row, column=3).value
            hour_type = ws.cell(row=row, column=9).value

            if not role:
                raise TimesheetError(
                    f"Missing role for '{employee}' in {path.name}, sheet '{sheet_name}', row {row}"
                )
            if not hour_type:
                raise TimesheetError(
                    f"Missing hour type for '{employee}' in {path.name}, sheet '{sheet_name}', row {row}"
                )

            hour_type = str(hour_type)
            if hour_type not in category_mapping:
                raise TimesheetError(
                    f"Unknown hour type '{hour_type}' in {path.name}, sheet '{sheet_name}', row {row}. "
                    f"Expected one of: {sorted(category_mapping)}"
                )

            category = category_mapping[hour_type]
            hours = row_total_hours(ws, row)
            if hours == 0:
                continue

            project_key = str(project) if project else NO_PROJECT_LABEL
            role_key = str(role)

            by_project[project_key][category] += hours
            by_role[role_key][category] += hours
            rows_processed += 1

    return by_project, by_role, rows_processed


def merge_aggregations(
    target: dict[str, dict[str, float]], source: dict[str, dict[str, float]]
) -> None:
    """Merge source aggregation into target in place."""
    for key, categories in source.items():
        for category, hours in categories.items():
            target[key][category] += hours


def category_total(categories: dict[str, float], category: str) -> float:
    return round(categories.get(category, 0.0), 1)


def row_total(categories: dict[str, float]) -> float:
    return round(sum(categories.get(cat, 0.0) for cat in REPORT_CATEGORIES), 1)


def write_report(
    output_path: Path,
    by_project: dict[str, dict[str, float]],
    by_role: dict[str, dict[str, float]],
    category_mapping: dict[str, str],
) -> None:
    """Write the monthly report workbook."""
    wb = openpyxl.Workbook()

    # Hours by Project
    ws_project = wb.active
    ws_project.title = "Hours by Project"
    project_headers = ["Project", *REPORT_CATEGORIES, "Total Hours"]
    ws_project.append(project_headers)

    def project_sort_key(project: str) -> tuple:
        if project == NO_PROJECT_LABEL:
            return (1, 0.0, project)
        return (0, -row_total(by_project[project]), project)

    for project in sorted(by_project.keys(), key=project_sort_key):
        cats = by_project[project]
        row = [project] + [category_total(cats, cat) for cat in REPORT_CATEGORIES]
        row.append(row_total(cats))
        ws_project.append(row)

    project_totals = defaultdict(float)
    for cats in by_project.values():
        for cat in REPORT_CATEGORIES:
            project_totals[cat] += cats.get(cat, 0.0)
    total_row = ["TOTAL"] + [round(project_totals[cat], 1) for cat in REPORT_CATEGORIES]
    total_row.append(round(sum(project_totals.values()), 1))
    ws_project.append(total_row)

    # Hours by Role
    ws_role = wb.create_sheet("Hours by Role")
    role_headers = ["Role", *REPORT_CATEGORIES, "Total Hours"]
    ws_role.append(role_headers)

    role_order = ["DevOps", "Developer", "QA Engineer", "Senior Developer"]
    sorted_roles = sorted(
        by_role.keys(),
        key=lambda r: (role_order.index(r) if r in role_order else len(role_order), r),
    )
    for role in sorted_roles:
        cats = by_role[role]
        row = [role] + [category_total(cats, cat) for cat in REPORT_CATEGORIES]
        row.append(row_total(cats))
        ws_role.append(row)

    role_totals = defaultdict(float)
    for cats in by_role.values():
        for cat in REPORT_CATEGORIES:
            role_totals[cat] += cats.get(cat, 0.0)
    total_row = ["TOTAL"] + [round(role_totals[cat], 1) for cat in REPORT_CATEGORIES]
    total_row.append(round(sum(role_totals.values()), 1))
    ws_role.append(total_row)

    # Category Mapping
    ws_mapping = wb.create_sheet("Category Mapping")
    ws_mapping.append(["Source Hour Type", "Report Category"])
    for source, target in sorted(category_mapping.items()):
        ws_mapping.append([source, target])

    wb.save(output_path)


def find_template(folder: Path) -> Path | None:
    """Locate the monthly report template in the folder, if present."""
    candidates = list(folder.glob("*Template*.xlsx"))
    return candidates[0] if candidates else None


def compare_with_reference(output_path: Path, reference_path: Path) -> list[str]:
    """Return a list of differences between output and reference workbooks."""
    if not reference_path.exists():
        return []

    out_wb = openpyxl.load_workbook(output_path, data_only=True)
    ref_wb = openpyxl.load_workbook(reference_path, data_only=True)
    differences: list[str] = []

    for sheet_name in ("Hours by Project", "Hours by Role"):
        if sheet_name not in out_wb.sheetnames or sheet_name not in ref_wb.sheetnames:
            continue
        out_ws = out_wb[sheet_name]
        ref_ws = ref_wb[sheet_name]

        out_rows = [
            [cell.value for cell in row]
            for row in out_ws.iter_rows(min_row=1, max_row=out_ws.max_row)
        ]
        ref_rows = [
            [cell.value for cell in row]
            for row in ref_ws.iter_rows(min_row=1, max_row=ref_ws.max_row)
        ]

        if len(out_rows) != len(ref_rows):
            differences.append(f"{sheet_name}: row count {len(out_rows)} vs {len(ref_rows)}")
            continue

        for i, (out_row, ref_row) in enumerate(zip(out_rows, ref_rows), start=1):
            if out_row != ref_row:
                differences.append(f"{sheet_name} row {i}: {out_row} != {ref_row}")

    return differences


def aggregate_timesheets(folder: Path) -> Path:
    """Process all team timesheets in folder and write Monthly_Report.xlsx."""
    if not folder.is_dir():
        raise TimesheetError(f"Folder not found: {folder}")

    timesheet_files = sorted(p for p in folder.iterdir() if is_timesheet_file(p))
    if not timesheet_files:
        raise TimesheetError(f"No team timesheet .xlsx files found in {folder}")

    template_path = find_template(folder)
    category_mapping = load_category_mapping(template_path)

    combined_by_project: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    combined_by_role: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    total_rows = 0

    print(f"Processing {len(timesheet_files)} timesheet file(s) from: {folder}")
    print(f"Category mapping: {dict(sorted(category_mapping.items()))}")
    print()

    for path in timesheet_files:
        by_project, by_role, rows = read_timesheet_file(path, category_mapping)
        merge_aggregations(combined_by_project, by_project)
        merge_aggregations(combined_by_role, by_role)
        total_rows += rows
        print(f"  {path.name}: {rows} row(s) processed")

    output_path = folder / OUTPUT_FILENAME
    write_report(output_path, combined_by_project, combined_by_role, category_mapping)

    print()
    print(f"Output written to: {output_path}")
    print(f"  Projects: {len(combined_by_project)}")
    print(f"  Roles:    {len(combined_by_role)}")
    print(f"  Total hours (all categories): {round(sum(sum(c.values()) for c in combined_by_project.values()), 1)}")

    reference_candidates = list(folder.glob("*REFERENCE*.xlsx"))
    if reference_candidates:
        diffs = compare_with_reference(output_path, reference_candidates[0])
        if diffs:
            print()
            print(f"Warning: output differs from reference ({reference_candidates[0].name}):")
            for diff in diffs[:10]:
                print(f"  - {diff}")
            if len(diffs) > 10:
                print(f"  ... and {len(diffs) - 10} more difference(s)")
        else:
            print()
            print(f"Output matches reference: {reference_candidates[0].name}")

    return output_path


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Aggregate weekly team timesheets into a monthly HR report."
    )
    parser.add_argument(
        "folder",
        type=Path,
        help="Folder containing team timesheet .xlsx files",
    )
    args = parser.parse_args()

    try:
        aggregate_timesheets(args.folder.resolve())
    except TimesheetError as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1
    except Exception as exc:
        print(f"Unexpected error: {exc}", file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    sys.exit(main())
